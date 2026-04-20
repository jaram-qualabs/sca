#!/usr/bin/env python3
"""
Simple Excel builder for checklist
"""
import os
import shutil
from openpyxl import load_workbook

src = '/Users/javieraramberri/Projects/SCA/Correccion/Checklist corrección.xlsx'
dst = '/Users/javieraramberri/Projects/SCA/sca-corrector/evals/runs/iteration-7/eval-1/with_skill/outputs/Checklist_Prueba1.xlsx'

os.makedirs(os.path.dirname(dst), exist_ok=True)
shutil.copy(src, dst)

wb = load_workbook(dst)
template_sheet = wb['Template Backend']

# Create new sheet from template
new_sheet = wb.copy_worksheet(template_sheet)
new_sheet.title = 'Prueba 1'

# Move to front
for i, sheet in enumerate(wb.sheetnames):
    if sheet == 'Prueba 1':
        wb.move_sheet('Prueba 1', offset=-(i))
        break

# Fill scores in column A
scores_dict = {
    3: 1,   # F3: Explica cómo correr
    4: 0,   # F4: Versión tecnología
    5: 1,   # F5: Explica decisiones
    8: 0,   # F8: Output consistente
    9: 1,   # F9: Parametriza archivos
    10: 1,  # F10: No hardcodea nombres
    11: 1,  # F11: No hardcodea cantidad
    12: 1,  # F12: No hardcodea providers
    13: 0,  # F13: Imprime como letra
    16: 1,  # F16: Nomenclatura
    17: 1,  # F17: Comentarios
    18: 1,  # F18: Sin comentarios excesivos
    19: 1,  # F19: Convenciones
    20: 1,  # F20: Divide en funciones
    21: 0,  # F21: Repite código A
    22: 0,  # F22: Código duplicado
    23: 1,  # F23: Sin mal indentado
    24: 1,  # F24: Sin formato irregular
    25: 1,  # F25: Error handling
    28: 1,  # F28: Parte A correcta
    29: 1,  # F29: Parte B cubre módulos
    30: 1,  # F30: Busca set reducido
    31: 0,  # F31: Set mínimo
    34: 2,  # Nivel: 2=junior
}

for row, value in scores_dict.items():
    cell = new_sheet[f'A{row}']
    cell.value = value

# Fill text fields
new_sheet['A37'] = 'Código bien organizado con funciones claras'
new_sheet['A38'] = 'Manejo de errores en ambas partes'
new_sheet['A39'] = 'Parametrización correcta con argparse'

new_sheet['A42'] = 'Duplicación: get_args() e process_user_info() son idénticas'
new_sheet['A43'] = 'Output inconsistente: Parte A escribe JSON, Parte B imprime consola'
new_sheet['A44'] = 'No documenta versión de Python'
new_sheet['A45'] = 'Greedy no garantiza mínimo (puede devolver >4 usuarios)'

new_sheet['A47'] = 'Buena organización. Mejoras: (1) importar funciones comunes, (2) output consistente, (3) documentar Python, (4) evaluar si greedy alcanza 4 usuarios'

wb.save(dst)
print(f'Excel created: {dst}')
