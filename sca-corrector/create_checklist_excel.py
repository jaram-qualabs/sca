#!/usr/bin/env python3
import sys
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Paths
src = '/Users/javieraramberri/Projects/SCA/Correccion/Checklist corrección.xlsx'
dst = '/Users/javieraramberri/Projects/SCA/sca-corrector/evals/runs/iteration-7/eval-1/with_skill/outputs/Checklist_Prueba1.xlsx'

# Ensure output dir
os.makedirs(os.path.dirname(dst), exist_ok=True)

# Copy template
shutil.copy(src, dst)

# Load and create new sheet
wb = load_workbook(dst)
template = wb['Template Backend']
ws = wb.copy_worksheet(template)
ws.title = 'Prueba 1'

# Move new sheet to front (after removing old blank sheets if needed)
ws_index = wb.index(ws)
wb.move_sheet(ws.title, offset=-ws_index)

# Scores: row -> value
scores = {
    3: 1,   # Explica cómo correr
    4: 0,   # Documenta versión
    5: 1,   # Explica decisiones
    8: 0,   # Output inconsistente
    9: 1,   # Parametriza archivos
    10: 1,  # No hardcodea nombres
    11: 1,  # No hardcodea cantidad
    12: 1,  # No hardcodea providers
    13: 0,  # Imprime como en letra
    16: 1,  # Nomenclatura consistente
    17: 1,  # Comentarios adecuados
    18: 1,  # Sin comentarios excesivos
    19: 1,  # Convenciones Python
    20: 1,  # Divide en funciones
    21: 0,  # Repite código A
    22: 0,  # Código duplicado
    23: 1,  # Sin mal indentado
    24: 1,  # Sin formato irregular
    25: 1,  # Error handling
    28: 1,  # Parte A correcta
    29: 1,  # Parte B cubre módulos
    30: 1,  # Busca set reducido
    31: 0,  # Set mínimo (greedy no garantiza)
    34: 2,  # Nivel: junior (0=no_suf, 1=trainee, 2=junior, 3=semi_sr)
}

# Set scores in column A
for row, val in scores.items():
    ws[f'A{row}'] = val

# Aspectos destacados (A37-A40)
ws['A37'] = 'Código bien organizado con funciones claras y bien nombradas'
ws['A38'] = 'Manejo de errores presente en ambas partes (try/except)'
ws['A39'] = 'Parametrización correcta con argparse en ambas partes'
ws['A40'] = ''

# Notas de corrección (A42-A45)
ws['A42'] = 'Duplicación de código: get_args() e process_user_info() son idénticas'
ws['A43'] = 'Output inconsistente: Parte A escribe a JSON, Parte B imprime a consola'
ws['A44'] = 'No documenta versión de Python en README'
ws['A45'] = 'Algoritmo greedy no garantiza solución mínima (puede devolver >4 usuarios)'

# Feedback (A47)
ws['A47'] = 'Buena organización del código. Mejoras necesarias: (1) importar funciones comunes en lugar de duplicarlas, (2) mantener output consistente, (3) documentar versión de Python, (4) evaluar si greedy alcanza 4 usuarios o requiere backtracking'

# Save workbook
wb.save(dst)
print(f"✅ Excel creado: {dst}")

# Recalculate formulas
recalc_script = '/var/folders/n1/7vkblp1d4lj_hc9_t82w28680000gn/T/claude-hostloop-plugins/7d6eec81aa93e9de/skills/xlsx/scripts/recalc.py'
print(f"📊 Recalculando fórmulas...")
result = os.system(f'python3 "{recalc_script}" "{dst}" 2>/dev/null')
if result == 0:
    print(f"✅ Fórmulas recalculadas")
else:
    print(f"⚠️  Error en recálculo (código {result})")

print(f"\n✅ Evaluación completada")
