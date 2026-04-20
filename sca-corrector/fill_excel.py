#!/usr/bin/env python3
import sys
import os

# Inline execution of Excel creation
if __name__ == '__main__':
    try:
        import shutil
        from pathlib import Path

        src = '/Users/javieraramberri/Projects/SCA/Correccion/Checklist corrección.xlsx'
        dst_dir = '/Users/javieraramberri/Projects/SCA/sca-corrector/evals/runs/iteration-7/eval-1/with_skill/outputs'
        dst = os.path.join(dst_dir, 'Checklist_Prueba1.xlsx')

        os.makedirs(dst_dir, exist_ok=True)
        shutil.copy(src, dst)

        # Try to fill
        from openpyxl import load_workbook
        wb = load_workbook(dst)

        # Create new sheet
        if 'Template Backend' in wb.sheetnames:
            template = wb['Template Backend']
            ws = wb.copy_worksheet(template)
            ws.title = 'Prueba 1'

            # Move to front
            idx = wb.index(ws)
            if idx > 0:
                for _ in range(idx):
                    wb.move_sheet('Prueba 1', offset=-1)

            # Fill all scores
            for row in [3, 4, 5, 8, 9, 10, 11, 12, 13, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 28, 29, 30, 31]:
                ws[f'A{row}'] = 1

            # Specific zeros
            ws['A4'] = 0
            ws['A8'] = 0
            ws['A13'] = 0
            ws['A21'] = 0
            ws['A22'] = 0
            ws['A31'] = 0

            # Level
            ws['A34'] = 2

            # Text fields
            ws['A37'] = 'Código bien organizado'
            ws['A38'] = 'Manejo de errores'
            ws['A42'] = 'Código duplicado entre partes'
            ws['A43'] = 'Output inconsistente'
            ws['A44'] = 'Falta versión Python'
            ws['A47'] = 'Mejoras: importar comunes, output consistente, versión, minimalidad'

            wb.save(dst)
            print(f'SUCCESS: {dst}')
        else:
            print(f'ERROR: No Template Backend sheet')
    except Exception as e:
        print(f'ERROR: {e}')
        import traceback
        traceback.print_exc()
