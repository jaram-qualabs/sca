#!/usr/bin/env python3
import sys
import os
import shutil

# Copy template first
src = '/Users/javieraramberri/Projects/SCA/Correccion/Checklist corrección.xlsx'
dst = '/Users/javieraramberri/Projects/SCA/sca-corrector/evals/runs/iteration-7/eval-1/with_skill/outputs/Checklist_Prueba1.xlsx'

os.makedirs(os.path.dirname(dst), exist_ok=True)

# Copy the template
print(f"Copying {src}...")
shutil.copy(src, dst)
print(f"✅ Template copied to {dst}")

# Now modify it using openpyxl
try:
    from openpyxl import load_workbook
    print("Loading workbook...")
    wb = load_workbook(dst)

    # Get template and create new sheet
    print(f"Available sheets: {wb.sheetnames}")
    if 'Template Backend' in wb.sheetnames:
        template = wb['Template Backend']
        print("Creating new sheet from template...")
        ws = wb.copy_worksheet(template)
        ws.title = 'Prueba 1'

        # Move sheet to front
        print("Moving sheet to front...")
        idx = wb.index(ws)
        if idx > 0:
            wb.move_sheet(ws.title, offset=-(idx))

        # Populate scores
        scores = {
            3: 1, 4: 0, 5: 1, 8: 0, 9: 1, 10: 1, 11: 1, 12: 1, 13: 0,
            16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 0, 22: 0, 23: 1, 24: 1, 25: 1,
            28: 1, 29: 1, 30: 1, 31: 0, 34: 2
        }

        print("Writing scores...")
        for row, val in scores.items():
            ws[f'A{row}'] = val

        print("Writing text fields...")
        ws['A37'] = 'Código bien organizado con funciones claras'
        ws['A38'] = 'Manejo de errores en ambas partes'
        ws['A39'] = 'Parametrización correcta con argparse'
        ws['A40'] = ''

        ws['A42'] = 'Duplicación: get_args() e process_user_info() idénticas'
        ws['A43'] = 'Output inconsistente entre partes'
        ws['A44'] = 'Falta documentar versión Python'
        ws['A45'] = 'Greedy no garantiza set mínimo'

        ws['A47'] = 'Buena organización. Mejoras: (1) importar funciones comunes, (2) output consistente, (3) documentar versión, (4) evaluar minimalidad'

        print("Saving workbook...")
        wb.save(dst)
        print(f"✅ Excel saved: {dst}")
    else:
        print(f"❌ Template Backend sheet not found. Available: {wb.sheetnames}")

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n✅ Done!")
